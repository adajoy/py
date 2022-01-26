import csv;
from openpyxl import load_workbook;

def getKeywordsByWs(ws):
    keywords = []
    keywordRange = ws['A']
    for cell in keywordRange:
        if cell.value is not None:
            keywords.append(cell.value)
    keywords.remove(keywords[0])
    return keywords

def getCreditDictByWs(ws):
    arr = []
    creditWordRange = ws['B']
    creditRange = ws['C']
    length = len(creditWordRange)
    for i in range(length):
        if creditWordRange[i].value is not None and creditRange[i].value is not None:
            arr.append((creditWordRange[i].value, creditRange[i].value))
    arr.remove(arr[0])
    return arr

def calcScore(keywords, creditDicts, comment):
    totalScore = 0
    matchCount = 0
    found = False
    for word in keywords:
        index = comment.find(word)
        if index >= 0:
            found = True
            start = index + len(word)
            substr = comment[start : start + 10]
            for tuple in creditDicts:
                if(substr.find(tuple[0]) > 0):
                    matchCount += 1
                    totalScore += tuple[1]
    avgScore = totalScore if matchCount == 0 else totalScore / matchCount
    return round(avgScore,2) if found else 'nan'

def writeCSV(comments):
    f = open('withCredit.csv', 'w', encoding='utf-8', newline='')
    csv_writer = csv.writer(f)
    csv_writer.writerow(['listing_id', 'location', 'date', 'comments', '总体评价', '描述相符', '办理入住', '性格特质', '服务态度', '沟通交流', '房源特色和环境', '房内设施', '交通便利指数', '他人推荐'])
    rows = []
    for obj in comments:
        valueList = []
        for key in obj.keys():
            valueList.append(obj[key])
        rows.append(valueList)
    csv_writer.writerows(rows)
    f.close()
    return

wb = load_workbook(filename='credit_lib.xlsx')

with open('filtered.csv', newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    comments = []
    for row in reader:
        comment = row['comments']
        scores = []
        for sheetname in wb.sheetnames:
            if sheetname != 'loy他人推荐':
                ws = wb[sheetname]
                keywords = getKeywordsByWs(ws)
                creditDicts = getCreditDictByWs(ws)
                score = calcScore(keywords, creditDicts, comment)
                scores.append(score)
                row[sheetname] = score
        comments.append(row)
        if len(comments) % 1000 == 0:
            print(len(comments))
    wsLast = wb['loy他人推荐']
    keyWords = []
    scores = []
    for word in wsLast['A']:
        if word.value is not None:
            keyWords.append(word.value)
    keyWords.remove(keyWords[0])
    for score in wsLast['B']:
        if score.value is not None:
            scores.append(score.value)
    scores.remove(scores[0])
    scoreMap = list(zip(keyWords, scores))
    for row in comments:
        totalScore = 0
        matchCount = 0
        comment = row['comments']
        for scoreTuple in scoreMap:
            if comment.find(scoreTuple[0]) >= 0:
                matchCount += 1
                totalScore += scoreTuple[1]
        avgScore = 'nan' if matchCount == 0 else totalScore / matchCount
        row['loy他人推荐'] = avgScore if isinstance(
            avgScore, str) else round(avgScore, 2)
    writeCSV(comments)
