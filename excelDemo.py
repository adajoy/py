from openpyxl import load_workbook
wb = load_workbook(filename = 'example.xlsx')
ws = wb["Sheet1"]
print(wb.sheetnames)
arr = []
# for col in ws.values:
#   for cell in col:
#     print(cell)
names = []
for cell in ws['A2': 'A4']:
  names.append(cell[0].value)

counts = []
for cell in ws['B2': 'B4']:
  counts.append(cell[0].value)

data = list(zip(names, counts))

def moreThan6(product):
  return product[1] > 6

filtered = filter(moreThan6, data)

print(list(filtered))