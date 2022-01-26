import csv
from geopy import distance
from openpyxl import Workbook
import copy
import re


def calcNeighbourScore(rows, key):
  if len(rows) == 0:
    return 'nan'
  sum = 0
  for row in rows:
    sum += float(row[key])
  return sum / len(rows)


with open('neighbourAvg.csv', newline='', encoding='utf-8') as f:
  reader = csv.DictReader(f)
  arr = []
  dimensions = []
  for i, row in enumerate(reader):
    if i == 0:
        dimensions = list(filter(lambda x: re.search(
            '[\u4e00-\u9fa5]', x), row.keys()))
    # if i > 100:
    #   break
    arr.append(row)
  arr1 = copy.deepcopy(arr)

  for i, row in enumerate(arr):
    row['neighbour'] = []
    row['sameYearNeighbor'] = []
    print(i)
    for j, row1 in enumerate(arr1):
        if i == j:
          continue
        if abs(float(row['latitude']) - float(row1['latitude'])) > 0.1:
          continue
        if abs(float(row['longitude']) - float(row1['longitude'])) > 0.1:
          continue
        coords_1 = (row['latitude'], row['longitude'])
        coords_2 = (row1['latitude'], row1['longitude'])
        d = distance.distance(coords_1, coords_2)
        if d < 0.5 and d > 0:
          row['neighbour'].append(copy.deepcopy(row1))
          if row1['year'] == row['year']:
            row['sameYearNeighbor'].append(copy.deepcopy(row1))

  wb = Workbook()
  for dimension in dimensions:
    ws = wb.create_sheet(dimension)
    ws.append(['listing_id', 'center_score', 'neighbor_score'])
    for i, row in enumerate(arr):
      ws.append([row['listing_id'], row[dimension],
                calcNeighbourScore(row['neighbour'], dimension)])
  wb.remove(wb.active)
  wb.save('wb.xlsx')

  wb1 = Workbook()
  for dimension in dimensions:
    ws = wb1.create_sheet(dimension)
    ws.append(['listing_id', 'year', 'center_score', 'neighbor_score'])
    for i, row in enumerate(arr):
      ws.append([row['listing_id'], row['year'], row[dimension],
                calcNeighbourScore(row['sameYearNeighbor'], dimension)])
  wb1.remove(wb1.active)
  wb1.save('wb_same_year.xlsx')
